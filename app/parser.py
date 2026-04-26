import os
import warnings

import numpy as np
import pandas as pd

warnings.simplefilter(action="ignore", category=FutureWarning)
warnings.simplefilter(action="ignore", category=UserWarning)


class TableParser:
    def parse_file(self, file_path: str) -> dict:
        _, ext = os.path.splitext(file_path)
        ext = ext.lower()
        if ext == ".xlsx":
            return self._parse_excel(file_path)
        elif ext == ".csv":
            return self._parse_csv(file_path)
        raise ValueError(f"Неподдерживаемый формат: {ext}")

    def _detect_type(self, series: pd.Series) -> str:
        s = series.dropna()
        if s.empty:
            return "empty"

        sample = s.head(500)
        n = len(sample)
        if n == 0:
            return "empty"

        bool_vals = {"true", "false", "1", "0", "yes", "no", "да", "нет"}
        if sample.astype(str).str.lower().isin(bool_vals).all():
            return "bool"

        try:
            if pd.to_numeric(sample, errors="coerce").notna().sum() / n > 0.8:
                return "number"
        except Exception:
            pass

        try:
            dt = pd.to_datetime(sample, errors="coerce")
            if dt.notna().sum() / n > 0.8:
                if sample.astype(str).str.contains(":").any():
                    return "datetime"
                return "date"
        except Exception:
            pass

        return "string"

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df

        df = df.dropna(how="all").dropna(axis=1, how="all")

        header_idx = 0
        max_density = 0

        for i in range(min(20, len(df))):
            row = df.iloc[i]
            density = row.notna().sum() / len(df.columns)
            if density > max_density and density > 0.25:
                row_str = " ".join(str(v) for v in row if pd.notna(v))
                if len(row_str) > 5:
                    max_density = density
                    header_idx = i

        new_header = df.iloc[header_idx]
        body = df.iloc[header_idx + 1 :].reset_index(drop=True)
        body.columns = new_header

        body.columns = [
            f"Col_{i}"
            if str(c).strip().startswith("Unnamed")
            or pd.isna(c)
            or str(c).strip() == ""
            else str(c).strip()
            for i, c in enumerate(body.columns)
        ]

        body = body.dropna(axis=1, how="all")

        body = body.ffill()

        body = body.replace([np.nan, np.inf, -np.inf, None], "", regex=True)

        body = body[body.astype(str).apply(lambda x: x.str.strip()).any(axis=1)]

        return body

    def _parse_excel(self, path: str) -> dict:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        result = {
            "file_type": "xlsx",
            "filename": os.path.basename(path),
            "sheets": {},
            "total_sheets": len(wb.sheetnames),
        }
        try:
            for name in wb.sheetnames:
                df_raw = pd.read_excel(path, sheet_name=name, header=None)
                if df_raw.empty:
                    continue

                df_clean = self._clean_dataframe(df_raw)
                if df_clean.empty:
                    continue

                cols_meta = [
                    {"name": str(c), "index": i, "type": self._detect_type(df_clean[c])}
                    for i, c in enumerate(df_clean.columns)
                ]

                r, c = df_clean.shape
                result["sheets"][name] = {
                    "sheet_name": name,
                    "columns": cols_meta,
                    "row_count": r,
                    "column_count": c,
                    "header_rows": 1,
                    "source_ref": {
                        "sheet": name,
                        "range": f"A1:{self._num2col(c)}{r}",
                        "row_start": 1,
                        "row_end": r,
                    },
                    "sample_data": df_clean.head(5).to_dict("records"),
                }
        finally:
            wb.close()
        return result

    def _parse_csv(self, path: str) -> dict:
        df_raw = None
        for enc in ["utf-8", "windows-1251", "cp1252"]:
            for sep in [",", ";", "\t"]:
                try:
                    df_raw = pd.read_csv(
                        path, encoding=enc, sep=sep, on_bad_lines="skip", header=None
                    )
                    if not df_raw.empty and len(df_raw.columns) > 1:
                        break
                except Exception:
                    continue
            if df_raw is not None and not df_raw.empty:
                break

        if df_raw is None or df_raw.empty:
            raise ValueError("Не удалось прочитать CSV файл")

        df_clean = self._clean_dataframe(df_raw)
        cols_meta = [
            {"name": str(c), "index": i, "type": self._detect_type(df_clean[c])}
            for i, c in enumerate(df_clean.columns)
        ]
        r, c = df_clean.shape

        return {
            "file_type": "csv",
            "filename": os.path.basename(path),
            "sheets": {
                "default_sheet": {
                    "sheet_name": "default_sheet",
                    "columns": cols_meta,
                    "row_count": r,
                    "column_count": c,
                    "header_rows": 1,
                    "source_ref": {
                        "sheet": "default_sheet",
                        "range": f"A1:{self._num2col(c)}{r}",
                        "row_start": 1,
                        "row_end": r,
                    },
                    "sample_data": df_clean.head(5).to_dict("records"),
                }
            },
            "total_sheets": 1,
        }

    @staticmethod
    def _num2col(n: int) -> str:
        res = ""
        while n > 0:
            n, rem = divmod(n - 1, 26)
            res = chr(65 + rem) + res
        return res
